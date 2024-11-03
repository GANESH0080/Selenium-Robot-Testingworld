import openpyxl

wk = openpyxl.load_workbook("D:\\RobotFramework\\AutomationUdemy2\\ExcelRead.xlsx")
sheetname = wk.sheetnames
print(sheetname[1])

sh=wk['Reading']
print(sh.max_row)
print(sh.max_column)

cell = sh.cell(1,1)
print(cell.value)

cell1 = sh.cell(1,2)
print(cell1.value)

print ("_____________________________________________________")

def fetch_number_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row

def fetch_cell_data(sheetname,row,cell):
    sh = wk[sheetname]
    ce = sh.cell(int(row), int(cell))
    return ce.value

print("_____________________Calling Calling Method Methods _______________________")


print(fetch_number_of_rows('Reading'))
print(fetch_cell_data('Reading',1,1))